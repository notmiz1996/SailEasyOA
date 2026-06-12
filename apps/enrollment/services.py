# apps/enrollment/services.py

"""
学员报名服务层

### EnrollmentService
- create_enrollment()       — 单条代录（查找/创建 Person + 创建 Enrollment）
- check_duplicate()         — 检查同一班级重复报名

### ImportService
- parse_excel()             — 解析 Excel 文件，返回行数据列表
- validate_row()            — 单行数据校验
- process_single_row()      — 单行处理（创建 Person + Enrollment）
- run_import()              — 批量导入入口（供 Django-Q 任务调用）
"""

"""
学员报名业务逻辑

批量导入服务（ImportService）：
- 支持带标题行的业务 Excel 格式（第2行为表头，自动识别）
- 列名别名映射（身份证号码 → 身份证号，电话 → 手机号）
- 忽略"序号"列和末尾无数据行（公司名/日期）
- 从地址字符串智能解析省市区（利用 Region 表数据，逐级匹配）
"""

import os
import re
from loguru import logger
from django.db import transaction

from apps.persons.models import Person, Region
from apps.training.models import TrainingClass
from .models import Enrollment, ImportErrorLog


class EnrollmentService:
    """学员报名服务"""

    @staticmethod
    def check_duplicate(person, training_class) -> bool:
        """检查同一人是否已在同一班级报名"""
        return Enrollment.objects.filter(
            person=person,
            training_class=training_class,
        ).exists()

    @staticmethod
    def create_enrollment(training_class, person_data: dict):
        """
        创建报名记录（查找或创建 Person）

        :param training_class: TrainingClass 实例
        :param person_data: {
            'name': str,
            'id_card': str,
            'phone': str (optional),
            'gender': str (optional),
            ...
        }
        :returns: Enrollment 实例
        :raises: ValidationError（重复报名时抛出）
        """
        id_card = person_data.pop('id_card')

        person, _ = Person.objects.get_or_create(
            id_card=id_card,
            defaults=person_data,
        )

        if person_data.get('name') and not person.name:
            person.name = person_data['name']
            person.save(update_fields=['name'])

        if EnrollmentService.check_duplicate(person, training_class):
            raise ValidationError('该学员已在本班级报名')

        enrollment = Enrollment.objects.create(
            person=person,
            training_class=training_class,
        )
        logger.info(f'学员 [{person.name}] 报名班级 [{training_class.name}] 成功')
        return enrollment



# ============================================================
# 列名别名映射（兼容不同数据源的 Excel 表头）
# ============================================================
COLUMN_ALIASES = {
    # 身份证
    '身份证号码': '身份证号',
    '身份证': '身份证号',
    # 地址
    '地址': '详细地址',
    '家庭地址': '详细地址',
    '通讯地址': '详细地址',
    # 电话
    '电话': '手机号',
    '联系电话': '手机号',
    '联系方式': '手机号',
    '手机号码': '手机号',
}

# 表头识别候选词（用于在多行中定位真正的表头行）
HEADER_CANDIDATES = {'姓名', '身份证号码', '身份证号', '性别', '电话', '地址', '序号', '备注'}

# 需忽略的列（读入数据时跳过）
IGNORED_COLUMNS = {'序号'}


class ImportService:
    """批量导入服务"""

    REQUIRED_HEADERS = ['姓名', '身份证号']
    OPTIONAL_HEADERS = ['手机号', '性别', '省', '市', '区', '详细地址', '紧急联系人', '紧急电话']

    # ----------------------------------------------------------------
    # 区域缓存
    # ----------------------------------------------------------------
    @staticmethod
    def _build_region_cache() -> dict:
        """
        预加载 Region 数据到内存，供地址解析使用。

        :returns: {
            'level3': {'封开县': Region, '封开': Region, ...},
            'level2': {'肇庆市': Region, '肇庆': Region, ...},
            'level1': {'广东省': Region, '广东': Region, ...},
        }
        每个 level 都包含带行政区划后缀和不带后缀两种 key。
        """
        cache = {'level1': {}, 'level2': {}, 'level3': {}}
        for region in Region.objects.all().select_related('parent'):
            key = region.name
            if region.level == 1:
                cache['level1'][key] = region
                # 省后缀：省、自治区、特别行政区
                for suffix in ('省', '市', '自治区', '特别行政区'):
                    if key.endswith(suffix):
                        cache['level1'][key[:-len(suffix)]] = region
            elif region.level == 2:
                cache['level2'][key] = region
                # 市后缀：市、自治州、地区、盟
                for suffix in ('市', '自治州', '地区', '盟'):
                    if key.endswith(suffix):
                        cache['level2'][key[:-len(suffix)]] = region
            elif region.level == 3:
                cache['level3'][key] = region
                # 区县后缀：区、县、县级市、自治县、旗、自治旗、特区、林区
                for suffix in ('区', '县', '县级市', '自治县', '旗', '自治旗', '特区', '林区'):
                    if key.endswith(suffix):
                        cache['level3'][key[:-len(suffix)]] = region
        return cache

    # ----------------------------------------------------------------
    # 地址智能解析
    # ----------------------------------------------------------------
    @staticmethod
    def _strip_name(text: str, name: str) -> str:
        """从文本中移除第一个匹配的名称，并清理分隔符。"""
        if name in text:
            text = text.replace(name, '', 1).strip()
            text = re.sub(r'^[\s\-–—,，、/]+', '', text)
        return text

    @staticmethod
    def _parse_address_regions(address: str, region_cache: dict) -> tuple:
        """
        从地址字符串中智能匹配省市区。

        策略：先匹配最细粒度（区/县，level 3），
        匹配成功后通过 parent 外键反查上级（省/市），
        避免对地址字符串的多次扫描。

        :param address: 原始地址字符串，如 "广东省封开县大洲镇大洲船队宿舍"
        :param region_cache: _build_region_cache() 的返回值
        :returns: (province_id, city_id, district_id, remaining_address)
                  remaining_address 是移除了省市区名称后的详细地址
        """
        if not address:
            return None, None, None, ''

        province = None
        city = None
        district = None
        remaining = address

        # --- Step 1: 匹配区/县（level 3） ---
        # 从最长名称开始匹配，避免 "封开" 先于 "封开县" 被匹配
        for name in sorted(region_cache['level3'], key=len, reverse=True):
            if name in remaining:
                district = region_cache['level3'][name]
                remaining = ImportService._strip_name(remaining, name)
                break

        # --- Step 2: 匹配市（level 2） ---
        if district:
            # 已匹配区级，通过 parent 反查市级
            city = district.parent
            if city:
                # 尝试从 remaining 中移除市名（带/不带后缀两种试探）
                remaining = ImportService._strip_name(remaining, city.name)
                for suffix in ('市', '自治州', '地区', '盟'):
                    if city.name.endswith(suffix):
                        remaining = ImportService._strip_name(
                            remaining, city.name[:-len(suffix)]
                        )
                        break
        else:
            for name in sorted(region_cache['level2'], key=len, reverse=True):
                if name in remaining:
                    city = region_cache['level2'][name]
                    remaining = ImportService._strip_name(remaining, name)
                    break

        # --- Step 3: 匹配省（level 1） ---
        if city:
            province = city.parent
        elif district:
            province = district.parent.parent if district.parent else None
        else:
            for name in sorted(region_cache['level1'], key=len, reverse=True):
                if name in remaining:
                    province = region_cache['level1'][name]
                    remaining = ImportService._strip_name(remaining, name)
                    break

        # 如果省是通过反查得到的，也从 remaining 中移除省名
        if province:
            remaining = ImportService._strip_name(remaining, province.name)
            for suffix in ('省', '市', '自治区', '特别行政区'):
                if province.name.endswith(suffix):
                    remaining = ImportService._strip_name(
                        remaining, province.name[:-len(suffix)]
                    )
                    break

        # --- 最终清理多余的空白和分隔符 ---
        remaining = re.sub(r'^[\s\-–—,，、/]+', '', remaining).strip()

        return (
            province.id if province else None,
            city.id if city else None,
            district.id if district else None,
            remaining,
        )

    # ----------------------------------------------------------------
    # Excel 解析（多行表头兼容）
    # ----------------------------------------------------------------
    @staticmethod
    def _normalize_headers(raw_headers: list) -> list:
        """
        将原始 Excel 表头统一规范化为内部标准列名。

        :param raw_headers: Excel 实际列名列表
        :returns: 规范化后的列名列表（未知列名保持原样）
        """
        normalized = []
        for h in raw_headers:
            if h is None:
                normalized.append(None)
                continue
            h = str(h).strip()
            # 取别名映射，无映射时保持原样
            mapped = COLUMN_ALIASES.get(h, h)
            normalized.append(mapped)
        return normalized

    @staticmethod
    def _find_header_row(ws, max_scan: int = 10) -> tuple:
        """
        在工作表中找到真正的表头行（跳过标题行）。

        扫描前 max_scan 行，找到包含最多 HEADER_CANDIDATES 的行视为表头。
        这样即使遇到：
          第1行：NPKB-...期二类轮机员培训学员名单表  ← 标题（分数低）
          第2行：序号 姓名 性别 身份证号码 地址 ...   ← 表头（分数高）
        也能正确识别。

        :param ws: openpyxl Worksheet
        :param max_scan: 最多扫描的行数
        :returns: (header_row_index, normalized_headers)
                  header_row_index 从 1 开始（Excel 行号）
        """
        best_row = 0
        best_score = 0
        best_headers = []

        for row_idx in range(1, min(max_scan, ws.max_row or max_scan) + 1):
            row_cells = [cell.value for cell in ws[row_idx]]
            if not row_cells or all(c is None for c in row_cells):
                continue

            score = 0
            for cell in row_cells:
                if cell and str(cell).strip() in HEADER_CANDIDATES:
                    score += 1

            if score > best_score:
                best_score = score
                best_row = row_idx
                best_headers = row_cells

        if best_score < 2:
            raise ValueError(
                '未找到合法的表头行（需要至少包含"姓名""身份证号码"等列名）'
            )

        normalized = ImportService._normalize_headers(best_headers)
        return best_row, normalized

    @staticmethod
    def parse_excel(file_path: str) -> list[dict]:
        """
        解析 Excel 文件，返回行数据列表。

        支持业务格式：
        - 第1行：表格标题（自动跳过）
        - 第N行：表头行（自动识别）
        - 后续行：数据
        - 末尾行：公司名称/制表日期（自动跳过——不含必填列数据）

        :param file_path: Excel 文件路径
        :returns: [{ '姓名': '...', '身份证号': '...', ... }, ...]
        :raises: ValueError（格式错误时抛出）
        """
        try:
            import openpyxl
        except ImportError:
            raise ValueError('缺少 openpyxl 库，请安装：pip install openpyxl')

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        if ws is None or (ws.max_row or 0) < 2:
            raise ValueError('文件为空或行数不足')

        # Step 1: 定位表头行（自动跳过标题行）
        header_row, headers = ImportService._find_header_row(ws)

        # Step 2: 校验必填列（使用规范化后的列名）
        for required in ImportService.REQUIRED_HEADERS:
            if required not in headers:
                raise ValueError(
                    f'缺少必填列：[{required}]'
                    f'（可接受的列名有：{list(COLUMN_ALIASES.keys())}）'
                )

        # Step 3: 建立列索引，跳过 IGNORED_COLUMNS（如"序号"）
        col_indices = {}
        for idx, h in enumerate(headers):
            if h is not None and h not in IGNORED_COLUMNS:
                col_indices[idx] = h

        # Step 4: 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_data = {}
            for idx, header_name in col_indices.items():
                if idx < len(row):
                    val = row[idx]
                    row_data[header_name] = str(val).strip() if val is not None else ''

            # 姓名和身份证号两者都必须有值（防止公司名/日期行误判）
            if all(row_data.get(h, '') for h in ImportService.REQUIRED_HEADERS):
                rows.append(row_data)

        wb.close()
        return rows

    # ----------------------------------------------------------------
    # 校验
    # ----------------------------------------------------------------
    @staticmethod
    def validate_row(row: dict, row_num: int, seen_id_cards: set) -> list[dict]:
        """单行数据校验。"""
        errors = []
        id_card = row.get('身份证号', '').strip()

        if not row.get('姓名', '').strip():
            errors.append({
                'row_num': row_num,
                'error_type': 'NAME_MISSING',
                'error_message': '姓名为空',
                'raw_data': row,
            })
            return errors

        if not id_card:
            errors.append({
                'row_num': row_num,
                'error_type': 'ID_CARD_MISSING',
                'error_message': '身份证号为空',
                'raw_data': row,
            })
            return errors

        if len(id_card) not in (15, 18):
            errors.append({
                'row_num': row_num,
                'error_type': 'ID_CARD_FORMAT',
                'error_message': f'身份证号长度非法：{len(id_card)}位（应为15或18位）',
                'raw_data': row,
            })
            return errors

        if id_card in seen_id_cards:
            errors.append({
                'row_num': row_num,
                'error_type': 'FILE_DUPLICATE',
                'error_message': f'文件内身份证号重复：{id_card}',
                'raw_data': row,
            })
            return errors

        seen_id_cards.add(id_card)
        return errors

    # ----------------------------------------------------------------
    # 单行处理（含地址解析）
    # ----------------------------------------------------------------
    @staticmethod
    def process_single_row(row: dict, training_class, region_cache: dict) -> tuple[bool, str]:
        """
        单行处理：创建 Person（含地址智能解析省市区）+ Enrollment。

        :param row: 行数据（列名已规范化）
        :param training_class: TrainingClass 实例
        :param region_cache: _build_region_cache() 的结果
        :returns: (success: bool, message: str)
        """
        try:
            with transaction.atomic():
                id_card = row['身份证号'].strip()
                name = row.get('姓名', '').strip()
                address_raw = row.get('详细地址', '').strip()
                phone = row.get('手机号', '').strip()
                gender = row.get('性别', '').strip()

                # 地址智能解析省市区
                province_id, city_id, district_id, address_detail = (
                    ImportService._parse_address_regions(address_raw, region_cache)
                )

                person, created = Person.objects.get_or_create(
                    id_card=id_card,
                    defaults={
                        'name': name,
                        'phone': phone,
                        'gender': gender,
                        'province_id': province_id,
                        'city_id': city_id,
                        'district_id': district_id,
                        'address_detail': address_detail,
                    },
                )

                if not created:
                    # 已存在的人员：只补充空字段，不覆盖已有数据
                    update_fields = []
                    if not person.name and name:
                        person.name = name
                        update_fields.append('name')
                    if not person.phone and phone:
                        person.phone = phone
                        update_fields.append('phone')
                    if not person.gender and gender:
                        person.gender = gender
                        update_fields.append('gender')
                    if not person.province_id and province_id:
                        person.province_id = province_id
                        update_fields.append('province_id')
                    if not person.city_id and city_id:
                        person.city_id = city_id
                        update_fields.append('city_id')
                    if not person.district_id and district_id:
                        person.district_id = district_id
                        update_fields.append('district_id')
                    if not person.address_detail and address_detail:
                        person.address_detail = address_detail
                        update_fields.append('address_detail')
                    if update_fields:
                        person.save(update_fields=update_fields)

                if Enrollment.objects.filter(
                    person=person, training_class=training_class
                ).exists():
                    return False, f'学员 [{name}] 已在本班级报名'

                Enrollment.objects.create(
                    person=person,
                    training_class=training_class,
                )
                return True, f'学员 [{name}] 报名成功'

        except Exception as e:
            return False, f'处理失败：{str(e)}'

    # ----------------------------------------------------------------
    # 批量导入入口
    # ----------------------------------------------------------------
    @staticmethod
    def run_import(class_id: str, file_path: str, task_id: str) -> dict:
        """
        批量导入入口（供 Django-Q 任务调用）。

        :returns: {
            'task_id': str,
            'total': int,
            'success': int,
            'fail': int,
            'errors': list[dict],
        }
        """
        training_class = TrainingClass.objects.get(pk=class_id)
        logger.info(f'批量导入开始：班级 [{training_class.name}]，文件 [{file_path}]')

        try:
            rows = ImportService.parse_excel(file_path)
        except ValueError as e:
            logger.error(f'批量导入文件解析失败：{e}')
            return {
                'task_id': task_id,
                'total': 0, 'success': 0, 'fail': 1,
                'errors': [{
                    'row_num': 0, 'error_type': 'FILE_PARSE_ERROR',
                    'error_message': str(e), 'raw_data': {},
                }],
            }

        # 预加载 Region 缓存（避免逐行查库，一次加载供所有行使用）
        region_cache = ImportService._build_region_cache()

        total = len(rows)
        success_count = 0
        fail_count = 0
        seen_id_cards = set()
        all_errors = []

        for idx, row in enumerate(rows):
            row_num = idx + 2

            errors = ImportService.validate_row(row, row_num, seen_id_cards)
            if errors:
                all_errors.extend(errors)
                fail_count += 1
                ImportErrorLog.objects.create(
                    task_id=task_id,
                    row_num=row_num,
                    error_type=errors[0]['error_type'],
                    error_message=errors[0]['error_message'],
                    raw_data=errors[0]['raw_data'],
                )
                continue

            ok, msg = ImportService.process_single_row(
                row, training_class, region_cache,
            )
            if ok:
                success_count += 1
            else:
                fail_count += 1
                error_entry = {
                    'row_num': row_num,
                    'error_type': 'PROCESS_FAIL',
                    'error_message': msg,
                    'raw_data': row,
                }
                all_errors.append(error_entry)
                ImportErrorLog.objects.create(
                    task_id=task_id,
                    row_num=row_num,
                    error_type='PROCESS_FAIL',
                    error_message=msg,
                    raw_data=row,
                )

        try:
            os.remove(file_path)
        except OSError:
            pass

        logger.info(
            f'批量导入完成：班级 [{training_class.name}]，'
            f'共 {total} 行，成功 {success_count}，失败 {fail_count}'
        )

        return {
            'task_id': task_id,
            'total': total,
            'success': success_count,
            'fail': fail_count,
            'errors': all_errors,
        }
# class ImportService:
#     """批量导入服务（支持从地址字符串智能识别省市区）"""
#
#     # 内部标准列名
#     REQUIRED_HEADERS = ['姓名', '身份证号']
#
#     # 列名别名映射：实际数据源列名 → 内部标准列名
#     COLUMN_ALIASES = {
#         '身份证号码': '身份证号',
#         '身份证': '身份证号',
#         '地址': '详细地址',
#         '家庭地址': '详细地址',
#         '通讯地址': '详细地址',
#         '电话': '手机号',
#         '联系电话': '手机号',
#         '联系方式': '手机号',
#         '手机号码': '手机号',
#     }
#
#     @staticmethod
#     def _normalize_headers(raw_headers: list) -> list:
#         """
#         将原始表头统一映射为内部标准列名
#
#         :param raw_headers: ['姓名', '身份证号码', '性别', '地址', '电话']
#         :returns: ['姓名', '身份证号', '性别', '详细地址', '手机号']
#         """
#         normalized = []
#         for h in raw_headers:
#             if h is None:
#                 normalized.append('')
#             else:
#                 h = h.strip()
#                 normalized.append(ImportService.COLUMN_ALIASES.get(h, h))
#         return normalized
#
#     @staticmethod
#     def _build_region_cache() -> dict:
#         """
#         预加载 Region 数据到内存，用于地址解析
#
#         数据结构：
#         {
#             'provinces': [(id, name, parent_id), ...],   # level=1
#             'cities': [(id, name, parent_id), ...],       # level=2
#             'districts': [(id, name, parent_id), ...],    # level=3
#         }
#         """
#         from apps.persons.models import Region
#
#         provinces = list(Region.objects.filter(level=1).values_list('id', 'name', 'parent_id'))
#         cities = list(Region.objects.filter(level=2).values_list('id', 'name', 'parent_id'))
#         districts = list(Region.objects.filter(level=3).values_list('id', 'name', 'parent_id'))
#
#         return {
#             'provinces': provinces,
#             'cities': cities,
#             'districts': districts,
#         }
#
#     @staticmethod
#     def _parse_address_regions(address: str, region_cache: dict) -> tuple:
#         """
#         从地址字符串中智能提取省市区（Region FK ID）
#
#         匹配策略：从最细粒度（区/县）开始，逐级向上。
#         示例：
#           "广东省封开县大洲镇..." → (广东ID, 肇庆ID, 封开县ID)
#           "广西桂平市蒙圩镇..."  → (广西ID, 桂平ID, None)  ← 区县级数据未加载时
#
#         :param address: 地址字符串
#         :param region_cache: _build_region_cache() 的返回值
#         :returns: (province_id, city_id, district_id) — 未匹配到则为 None
#         """
#         if not address:
#             return None, None, None
#
#         province_id = None
#         city_id = None
#         district_id = None
#
#         # ---- 1. 尝试匹配区/县（level 3）— 最精确 ----
#         for dist_id, dist_name, dist_parent_id in region_cache['districts']:
#             if dist_name in address:
#                 district_id = dist_id
#                 city_id = dist_parent_id
#                 # 根据 city_id 反查 province_id
#                 for c_id, c_name, c_parent_id in region_cache['cities']:
#                     if c_id == city_id:
#                         province_id = c_parent_id
#                         break
#                 break
#
#         # ---- 2. 没匹配到区/县，尝试匹配市（level 2）----
#         if district_id is None:
#             for c_id, c_name, c_parent_id in region_cache['cities']:
#                 if c_name in address:
#                     city_id = c_id
#                     province_id = c_parent_id
#                     break
#
#         # ---- 3. 还没匹配到市，尝试匹配省（level 1）----
#         if province_id is None:
#             for p_id, p_name, _ in region_cache['provinces']:
#                 if p_name in address:
#                     province_id = p_id
#                     break
#
#         return province_id, city_id, district_id
#
#     @staticmethod
#     def parse_excel(file_path: str) -> list[dict]:
#         """
#         解析 Excel 文件，返回行数据列表
#
#         支持列名别名映射，例如：
#         - 源列「身份证号码」→ 内部使用「身份证号」
#         - 源列「地址」→ 内部使用「详细地址」
#         - 源列「电话」→ 内部使用「手机号」
#
#         :param file_path: Excel 文件路径
#         :returns: [{ '姓名': '...', '身份证号': '...', ... }, ...]
#         :raises: ValueError（格式错误时抛出）
#         """
#         try:
#             import openpyxl
#         except ImportError:
#             raise ValueError('缺少 openpyxl 库，请安装：pip install openpyxl')
#
#         wb = openpyxl.load_workbook(file_path, read_only=True)
#         ws = wb.active
#         if ws is None or ws.max_row < 2:
#             raise ValueError('文件为空或只有表头')
#
#         # 读取并规范化表头
#         raw_headers = [cell.value for cell in ws[1]]
#         if not raw_headers or raw_headers[0] is None:
#             raise ValueError('文件格式错误：第一行必须为表头')
#
#         headers = ImportService._normalize_headers(raw_headers)
#
#         # 校验必填列
#         missing = [h for h in ImportService.REQUIRED_HEADERS if h not in headers]
#         if missing:
#             raise ValueError(f'缺少必填列：{missing}（可接受的列名：{missing} 或 身份证号码）')
#
#         # 逐行读取数据
#         rows = []
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             row_data = {}
#             for idx, header in enumerate(headers):
#                 if header and idx < len(row):
#                     val = row[idx]
#                     row_data[header] = str(val).strip() if val is not None else ''
#             # 跳过全空行
#             if any(row_data.get(h, '') for h in ImportService.REQUIRED_HEADERS):
#                 rows.append(row_data)
#
#         wb.close()
#         return rows
#
#     @staticmethod
#     def validate_row(row: dict, row_num: int, seen_id_cards: set, region_cache: dict = None) -> list[dict]:
#         """
#         单行数据校验
#
#         校验项：
#         - 姓名非空
#         - 身份证号格式（15或18位）
#         - 文件内重复
#         - 地址解析省市区（非必检，解析失败不阻塞）
#
#         :returns: 错误列表
#         """
#         errors = []
#         id_card = row.get('身份证号', '').strip()
#
#         if not row.get('姓名', '').strip():
#             errors.append({
#                 'row_num': row_num,
#                 'error_type': 'NAME_MISSING',
#                 'error_message': '姓名为空',
#                 'raw_data': row,
#             })
#             return errors
#
#         if not id_card:
#             errors.append({
#                 'row_num': row_num,
#                 'error_type': 'ID_CARD_MISSING',
#                 'error_message': '身份证号为空',
#                 'raw_data': row,
#             })
#             return errors
#
#         if len(id_card) not in (15, 18):
#             errors.append({
#                 'row_num': row_num,
#                 'error_type': 'ID_CARD_FORMAT',
#                 'error_message': f'身份证号长度非法：{len(id_card)}位（应为15或18位）',
#                 'raw_data': row,
#             })
#             return errors
#
#         if id_card in seen_id_cards:
#             errors.append({
#                 'row_num': row_num,
#                 'error_type': 'FILE_DUPLICATE',
#                 'error_message': f'文件内身份证号重复：{id_card}',
#                 'raw_data': row,
#             })
#             return errors
#
#         seen_id_cards.add(id_card)
#         return errors
#
#     @staticmethod
#     def process_single_row(row: dict, training_class, region_cache: dict = None) -> tuple[bool, str]:
#         """
#         单行处理：创建 Person + Enrollment
#
#         从 address_detail 中智能提取省市区（Region FK），
#         如果 Region 表中没有对应数据则留空，不阻塞导入。
#
#         :returns: (success: bool, message: str)
#         """
#         try:
#             with transaction.atomic():
#                 id_card = row['身份证号'].strip()
#                 name = row.get('姓名', '').strip()
#                 address = row.get('详细地址', '').strip()
#
#                 # ---- 从地址中智能提取省市区 ----
#                 province_id, city_id, district_id = None, None, None
#                 if address and region_cache:
#                     province_id, city_id, district_id = (
#                         ImportService._parse_address_regions(address, region_cache)
#                     )
#
#                 # ---- 查找或创建 Person ----
#                 person, created = Person.objects.get_or_create(
#                     id_card=id_card,
#                     defaults={
#                         'name': name,
#                         'phone': row.get('手机号', '').strip(),
#                         'gender': row.get('性别', '').strip(),
#                         'province_id': province_id,
#                         'city_id': city_id,
#                         'district_id': district_id,
#                         'address_detail': address,
#                     },
#                 )
#
#                 # 更新 Person 信息（如果用临时档案替换了真实姓名或地址）
#                 if not created:
#                     update_fields = []
#                     if not person.name and name:
#                         person.name = name
#                         update_fields.append('name')
#                     # 如果 Person 之前没有地址，现在补上
#                     if not person.address_detail and address:
#                         person.address_detail = address
#                         update_fields.append('address_detail')
#                     # 如果 Person 之前没有省市区，现在尝试补上
#                     if not person.province_id and province_id:
#                         person.province_id = province_id
#                         update_fields.append('province_id')
#                     if not person.city_id and city_id:
#                         person.city_id = city_id
#                         update_fields.append('city_id')
#                     if not person.district_id and district_id:
#                         person.district_id = district_id
#                         update_fields.append('district_id')
#                     if update_fields:
#                         person.save(update_fields=update_fields)
#
#                 # ---- 检查重复报名 ----
#                 if Enrollment.objects.filter(person=person, training_class=training_class).exists():
#                     return False, f'学员 [{name}] 已在本班级报名'
#
#                 # ---- 创建 Enrollment ----
#                 Enrollment.objects.create(
#                     person=person,
#                     training_class=training_class,
#                 )
#                 return True, f'学员 [{name}] 报名成功'
#
#         except Exception as e:
#             return False, f'处理失败：{str(e)}'
#
#     @staticmethod
#     def run_import(class_id: str, file_path: str, task_id: str) -> dict:
#         """
#         批量导入入口（供 Django-Q 任务调用）
#
#         :returns: {
#             'task_id': str,
#             'total': int,
#             'success': int,
#             'fail': int,
#             'errors': list[dict],
#         }
#         """
#         training_class = TrainingClass.objects.get(pk=class_id)
#         logger.info(f'批量导入开始：班级 [{training_class.name}]，文件 [{file_path}]')
#
#         # 解析 Excel
#         try:
#             rows = ImportService.parse_excel(file_path)
#         except ValueError as e:
#             logger.error(f'批量导入文件解析失败：{e}')
#             return {
#                 'task_id': task_id,
#                 'total': 0, 'success': 0, 'fail': 1,
#                 'errors': [
#                     {'row_num': 0, 'error_type': 'FILE_PARSE_ERROR', 'error_message': str(e), 'raw_data': {}}],
#             }
#
#         # 预加载 Region 数据（用于地址解析）
#         region_cache = ImportService._build_region_cache()
#         logger.info(
#             f'Region 缓存加载完成：{len(region_cache["provinces"])}省 {len(region_cache["cities"])}市 {len(region_cache["districts"])}区/县')
#
#         total = len(rows)
#         success_count = 0
#         fail_count = 0
#         seen_id_cards = set()
#         all_errors = []
#
#         for idx, row in enumerate(rows):
#             row_num = idx + 2
#
#             # 校验
#             errors = ImportService.validate_row(row, row_num, seen_id_cards, region_cache)
#             if errors:
#                 all_errors.extend(errors)
#                 fail_count += 1
#                 ImportErrorLog.objects.create(
#                     task_id=task_id,
#                     row_num=row_num,
#                     error_type=errors[0]['error_type'],
#                     error_message=errors[0]['error_message'],
#                     raw_data=errors[0]['raw_data'],
#                 )
#                 continue
#
#             # 处理（含地址智能解析）
#             ok, msg = ImportService.process_single_row(row, training_class, region_cache)
#             if ok:
#                 success_count += 1
#             else:
#                 fail_count += 1
#                 error_entry = {
#                     'row_num': row_num,
#                     'error_type': 'PROCESS_FAIL',
#                     'error_message': msg,
#                     'raw_data': row,
#                 }
#                 all_errors.append(error_entry)
#                 ImportErrorLog.objects.create(
#                     task_id=task_id,
#                     row_num=row_num,
#                     error_type='PROCESS_FAIL',
#                     error_message=msg,
#                     raw_data=row,
#                 )
#
#         # 清理临时文件
#         try:
#             os.remove(file_path)
#         except OSError:
#             pass
#
#         logger.info(
#             f'批量导入完成：班级 [{training_class.name}]，'
#             f'共 {total} 行，成功 {success_count}，失败 {fail_count}'
#         )
#
#         return {
#             'task_id': task_id,
#             'total': total,
#             'success': success_count,
#             'fail': fail_count,
#             'errors': all_errors,
#         }