# utils/excel.py

"""
Excel 导入导出工具
基于 openpyxl 实现
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path


def create_template(headers: list[str], sheet_name: str = 'Sheet1') -> Workbook:
    """
    创建标准 Excel 模板
    - 首行加粗、居中、浅蓝色背景
    - 自动列宽
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 自动列宽
    for col, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = len(header) * 2 + 4

    return wb


def read_excel(file_path: str, sheet_name: str = None) -> tuple[list[str], list[list]]:
    """
    读取 Excel 文件
    返回 (headers, rows)
    - headers：第一行表头列表
    - rows：从第二行开始的数据行列表
    """
    wb = load_workbook(file_path, read_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(cell) if cell is not None else '' for cell in rows[0]]
    data_rows = [list(row) for row in rows[1:] if any(cell is not None for cell in row)]

    wb.close()
    return headers, data_rows